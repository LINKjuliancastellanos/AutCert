"""
GENERADOR DE ARCHIVO DE CONTROL DE EJECUCION DE PRUEBAS - AutCert
=================================================================
Genera un archivo Excel con el estado de ejecucion de pruebas por HU,
tomando datos del CSV y analizando las carpetas de evidencias.

Columnas generadas:
- ID HU, Categoria, Tester Asignado (desde CSV)
- Total Test Cases, Test Cases Completos, Pendientes, Avance % (calculados)
- Estado Certificacion, Estado TCs, TCs Activos (automaticos)
- Estado Subida, Observaciones (manuales con validacion)

Uso: python generar_control.py
"""

import sys
import os
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# Configurar encoding para Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# Obtener la raiz del proyecto (3 niveles arriba: script -> Control_Ejecucion -> Modulos -> AutCert)
PROJECT_ROOT = Path(__file__).parent.parent.parent.resolve()

# Agregar al path para importar config_paths
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
except ImportError as e:
    print(f"Error: Falta dependencia {e}")
    print("Instalar con: pip install pandas openpyxl")
    sys.exit(1)

from config_paths import (
    CSV_PATH_STR,
    EVIDENCIAS_ROOT_STR,
    MODULOS_DIR
)

# Configuracion del modulo
CONTROL_DIR = MODULOS_DIR / "Control_Ejecucion"
OUTPUT_DIR = CONTROL_DIR / "datos"
OUTPUT_FILE = OUTPUT_DIR / "control_ejecucion.xlsx"

# Mapeo de categorias CSV -> carpetas de evidencias
CATEGORIA_MAPPING = {
    "Afiliación y Novedades": "AyN",
    "Afiliacion y Novedades": "AyN",
    "Recaudo y Cartera": "RyC",
    "Transversales": "Transversal",
    "Transversal": "Transversal"
}

# Colores para estados
COLORES = {
    # Estado Certificacion
    "Sin Test Cases": "D9D9D9",      # Gris
    "Sin Iniciar": "FF6B6B",          # Rojo
    "En Progreso": "FFE066",          # Amarillo
    "Completo": "69DB7C",             # Verde

    # Estado Subida
    "Por Subir": "4DABF7",            # Azul
    "Subido": "69DB7C",               # Verde
    "Bloqueado": "868E96",            # Gris
    "Preguntar": "FFE066",            # Amarillo
    "Sin Iniciar Subida": "FF6B6B",   # Rojo

    # Header
    "Header": "2C3E50",               # Azul oscuro
    "HeaderFont": "FFFFFF"            # Blanco
}


def verificar_archivo_cerrado():
    """
    Verifica si el archivo Excel esta cerrado y puede ser modificado.
    Si esta abierto, solicita al usuario que lo cierre.
    Returns: True si el archivo esta disponible, False si el usuario cancela.
    """
    if not OUTPUT_FILE.exists():
        return True  # No existe, se puede crear

    while True:
        try:
            # Intentar abrir el archivo en modo escritura para verificar si esta bloqueado
            with open(str(OUTPUT_FILE), 'a'):
                pass
            return True  # Archivo disponible
        except PermissionError:
            print("\n" + "="*60)
            print("ARCHIVO EN USO")
            print("="*60)
            print(f"\nEl archivo '{OUTPUT_FILE.name}' esta abierto en Excel.")
            print("Por favor, cierre el archivo para continuar con la actualizacion.")
            print("\nOpciones:")
            print("  [ENTER] - Reintentar despues de cerrar el archivo")
            print("  [C]     - Cancelar la actualizacion")
            print()

            respuesta = input("Presione ENTER para reintentar o 'C' para cancelar: ").strip().lower()

            if respuesta == 'c':
                print("\nActualizacion cancelada por el usuario.")
                return False

            print("\nVerificando nuevamente...")


def cargar_csv():
    """Carga el archivo CSV de datos"""
    if not os.path.exists(CSV_PATH_STR):
        print(f"Error: No se encontro el archivo CSV: {CSV_PATH_STR}")
        return None

    try:
        df = pd.read_csv(CSV_PATH_STR, encoding='utf-8-sig')
        print(f"CSV cargado: {len(df)} registros")
        return df
    except Exception as e:
        print(f"Error al cargar CSV: {e}")
        return None


def cargar_datos_manuales_existentes():
    """
    Carga los datos manuales (Estado Subida, Observaciones) del archivo existente.
    Returns: dict {hu_id: {'estado_subida': str, 'observaciones': str}}
    """
    datos_manuales = {}

    if not OUTPUT_FILE.exists():
        print("No existe archivo previo, se creara uno nuevo")
        return datos_manuales

    try:
        wb = load_workbook(str(OUTPUT_FILE), read_only=True, data_only=True)
        ws = wb.active

        # La fila 1 es el hipervinculo, fila 2 es header, datos desde fila 3
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]:  # Si hay ID HU
                hu_id = str(row[0])
                # Columna L (indice 11) = Estado Subida, Columna M (indice 12) = Observaciones
                estado_subida = row[11] if len(row) > 11 and row[11] else ""
                observaciones = row[12] if len(row) > 12 and row[12] else ""

                if estado_subida or observaciones:
                    datos_manuales[hu_id] = {
                        'estado_subida': estado_subida,
                        'observaciones': observaciones
                    }

        wb.close()
        print(f"Datos manuales recuperados: {len(datos_manuales)} HUs con datos")

    except Exception as e:
        print(f"Aviso: No se pudieron cargar datos previos: {e}")

    return datos_manuales


def construir_mapa_hu_tcs(df):
    """
    Construye un diccionario que mapea cada HU a sus Test Cases.
    El CSV tiene estructura secuencial: HU seguida de sus TCs.

    Returns:
        dict: {hu_id: {'info': {datos_hu}, 'tcs': [lista_de_tcs]}}
    """
    mapa = {}
    hu_actual = None

    for _, row in df.iterrows():
        work_type = row.get('Work Item Type', '')

        if work_type == 'User Story':
            hu_id = str(row['ID'])
            hu_actual = hu_id
            mapa[hu_id] = {
                'info': {
                    'id': hu_id,
                    'title': row.get('Title', ''),
                    'categoria': row.get('Modulo Balu', ''),
                    'estado_hu': row.get('State', ''),
                    'tester': row.get('Tester', ''),
                    'aprobador': row.get('Aprobador QA', '')
                },
                'tcs': []
            }

        elif work_type == 'Test Case' and hu_actual:
            tc_info = {
                'id': str(row['ID']),
                'title': row.get('Title', ''),
                'reason': row.get('Reason', ''),
                'state': row.get('State', '')
            }
            mapa[hu_actual]['tcs'].append(tc_info)

    print(f"HUs procesadas: {len(mapa)}")
    return mapa


def obtener_carpeta_categoria(categoria):
    """Convierte el nombre de categoria del CSV a nombre de carpeta"""
    if pd.isna(categoria) or str(categoria).strip() == '':
        return None

    categoria_str = str(categoria).strip()
    return CATEGORIA_MAPPING.get(categoria_str, None)


def contar_tcs_completados(hu_id, categoria):
    """
    Cuenta cuantos TCs tienen evidencias completadas.
    Una carpeta CP* con al menos 1 archivo = 1 TC completado.
    """
    carpeta_cat = obtener_carpeta_categoria(categoria)

    if carpeta_cat is None:
        # Buscar en todas las categorias
        for cat_folder in ["AyN", "RyC", "Transversal"]:
            hu_path = Path(EVIDENCIAS_ROOT_STR) / cat_folder / f"HU{hu_id}"
            if hu_path.exists():
                return _contar_carpetas_cp_con_archivos(hu_path)
        return 0

    hu_path = Path(EVIDENCIAS_ROOT_STR) / carpeta_cat / f"HU{hu_id}"

    if not hu_path.exists():
        return 0

    return _contar_carpetas_cp_con_archivos(hu_path)


def _contar_carpetas_cp_con_archivos(hu_path):
    """Cuenta carpetas CP* que contienen al menos un archivo"""
    count = 0

    try:
        for item in hu_path.iterdir():
            if item.is_dir() and item.name.startswith("CP"):
                # Verificar si tiene al menos un archivo
                archivos = list(item.glob("*"))
                if any(f.is_file() for f in archivos):
                    count += 1
    except Exception as e:
        print(f"  Advertencia: Error al leer {hu_path}: {e}")

    return count


def calcular_avance(completos, total):
    """Calcula el porcentaje de avance"""
    if total == 0:
        return 0.0
    return round((completos / total) * 100, 1)


def determinar_estado_certificacion(total, completos):
    """Determina el estado de certificacion segun las reglas"""
    if total == 0:
        return "Sin Test Cases"
    elif completos == 0:
        return "Sin Iniciar"
    elif completos < total:
        return "En Progreso"
    else:
        return "Completo"


def analizar_estado_tcs(tcs_list):
    """
    Analiza el estado de los TCs basado en el campo Reason.
    - Obsolete = TC obsoleto
    - Vacio o cualquier otro = TC activo
    """
    if not tcs_list:
        return "Sin Test Cases"

    total = len(tcs_list)
    obsoletos = sum(1 for tc in tcs_list if str(tc.get('reason', '')).strip().lower() == 'obsolete')
    activos = total - obsoletos

    if obsoletos == total:
        return "Todos obsoletos"
    elif obsoletos == 0:
        return "Todos activos"
    else:
        return f"Parcial ({activos}/{total} activos)"


def obtener_tcs_activos(tcs_list):
    """Obtiene la lista de IDs de TCs no obsoletos"""
    activos = []

    for tc in tcs_list:
        reason = str(tc.get('reason', '')).strip().lower()
        if reason != 'obsolete':
            activos.append(tc['id'])

    return activos


def generar_datos_control(mapa_hu_tcs, datos_manuales=None):
    """
    Genera los datos para el archivo de control.
    Si se proporcionan datos_manuales, conserva Estado Subida y Observaciones.
    """
    if datos_manuales is None:
        datos_manuales = {}

    datos = []

    print("\nProcesando HUs...")

    for hu_id, hu_data in mapa_hu_tcs.items():
        info = hu_data['info']
        tcs = hu_data['tcs']

        # Datos basicos
        categoria = info['categoria']
        estado_hu = info['estado_hu']
        tester = info['tester']

        # Calculos
        total_tcs = len(tcs)
        tcs_completos = contar_tcs_completados(hu_id, categoria)
        tcs_pendientes = max(0, total_tcs - tcs_completos)
        avance = calcular_avance(tcs_completos, total_tcs)

        # Estados
        estado_cert = determinar_estado_certificacion(total_tcs, tcs_completos)
        estado_tcs = analizar_estado_tcs(tcs)
        tcs_activos = obtener_tcs_activos(tcs)
        tcs_activos_str = ", ".join(tcs_activos) if tcs_activos else ""

        # Limpiar tester (quitar email si existe)
        if pd.notna(tester) and '<' in str(tester):
            tester = str(tester).split('<')[0].strip()

        # Recuperar datos manuales si existen
        manual = datos_manuales.get(hu_id, {})
        estado_subida = manual.get('estado_subida', "")
        observaciones = manual.get('observaciones', "")

        datos.append({
            'ID HU': hu_id,
            'Categoria': categoria if pd.notna(categoria) else "",
            'Estado HU': estado_hu if pd.notna(estado_hu) else "",
            'Tester Asignado': tester if pd.notna(tester) else "",
            'Total Test Cases': total_tcs,
            'Test Cases Completos': tcs_completos,
            'Test Cases Pendientes': tcs_pendientes,
            'Avance %': avance,
            'Estado Certificacion': estado_cert,
            'Estado TCs': estado_tcs,
            'TCs Activos (IDs)': tcs_activos_str,
            'Estado Subida': estado_subida,  # Conservado de archivo anterior
            'Observaciones': observaciones   # Conservado de archivo anterior
        })

    print(f"Datos generados para {len(datos)} HUs")
    return datos


def crear_archivo_excel(datos):
    """Crea el archivo Excel con formato y validaciones"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Control de Ejecucion"

    # Definir columnas
    columnas = [
        'ID HU', 'Categoria', 'Estado HU', 'Tester Asignado', 'Total Test Cases',
        'Test Cases Completos', 'Test Cases Pendientes', 'Avance %',
        'Estado Certificacion', 'Estado TCs', 'TCs Activos (IDs)',
        'Estado Subida', 'Observaciones'
    ]

    # Anchos de columna
    anchos = [10, 25, 28, 35, 15, 18, 18, 12, 20, 22, 40, 15, 40]

    # Estilos
    header_fill = PatternFill(start_color=COLORES["Header"], end_color=COLORES["Header"], fill_type="solid")
    header_font = Font(bold=True, color=COLORES["HeaderFont"], size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === FILA 1: Hipervinculo para actualizar ===
    # Combinar celdas A1:M1
    ws.merge_cells('A1:M1')
    btn_cell = ws.cell(row=1, column=1)

    # Ruta al archivo .bat
    bat_path = str(CONTROL_DIR / "Generar_Control.bat")
    btn_cell.hyperlink = bat_path
    btn_cell.value = "ACTUALIZAR ARCHIVO DE CONTROL (Click aqui)"
    btn_cell.font = Font(bold=True, color="FFFFFF", size=12, underline="single")
    btn_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    btn_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # === FILA 2: Headers ===
    for col_idx, (columna, ancho) in enumerate(zip(columnas, anchos), 1):
        cell = ws.cell(row=2, column=col_idx, value=columna)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # Altura del header
    ws.row_dimensions[2].height = 30

    # === FILAS 3+: Datos ===
    for row_idx, fila in enumerate(datos, 3):
        for col_idx, columna in enumerate(columnas, 1):
            valor = fila.get(columna, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Aplicar colores a Estado Certificacion (columna 9)
            if col_idx == 9 and valor in COLORES:
                cell.fill = PatternFill(start_color=COLORES[valor], end_color=COLORES[valor], fill_type="solid")
                if valor in ["Sin Iniciar", "Sin Test Cases"]:
                    cell.font = Font(color="FFFFFF")

            # Centrar columnas numericas
            if col_idx in [1, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar validacion de lista desplegable para Estado Subida (columna 12 = L)
    opciones_subida = '"Por Subir,Subido,Bloqueado,Preguntar,Sin Iniciar"'
    dv = DataValidation(type="list", formula1=opciones_subida, allow_blank=True)
    dv.error = "Seleccione una opcion de la lista"
    dv.errorTitle = "Valor invalido"
    dv.prompt = "Seleccione el estado de subida"
    dv.promptTitle = "Estado Subida"

    # Aplicar validacion a todas las filas de datos (fila 3 en adelante)
    if len(datos) > 0:
        ultima_fila = len(datos) + 2  # +2 porque datos empiezan en fila 3
        dv.add(f"L3:L{ultima_fila}")
        ws.add_data_validation(dv)

        # Agregar formato condicional para Estado Subida (columna L)
        rango_subida = f"L3:L{ultima_fila}"

        # Por Subir -> Azul
        ws.conditional_formatting.add(
            rango_subida,
            FormulaRule(
                formula=['$L3="Por Subir"'],
                fill=PatternFill(start_color="4DABF7", end_color="4DABF7", fill_type="solid")
            )
        )
        # Subido -> Verde
        ws.conditional_formatting.add(
            rango_subida,
            FormulaRule(
                formula=['$L3="Subido"'],
                fill=PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
            )
        )
        # Bloqueado -> Gris
        ws.conditional_formatting.add(
            rango_subida,
            FormulaRule(
                formula=['$L3="Bloqueado"'],
                fill=PatternFill(start_color="868E96", end_color="868E96", fill_type="solid"),
                font=Font(color="FFFFFF")
            )
        )
        # Preguntar -> Amarillo
        ws.conditional_formatting.add(
            rango_subida,
            FormulaRule(
                formula=['$L3="Preguntar"'],
                fill=PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
            )
        )
        # Sin Iniciar -> Rojo
        ws.conditional_formatting.add(
            rango_subida,
            FormulaRule(
                formula=['$L3="Sin Iniciar"'],
                fill=PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
                font=Font(color="FFFFFF")
            )
        )

    # Congelar panel (fila 1 hipervinculo + fila 2 header fijos)
    ws.freeze_panes = "A3"

    # Filtros automaticos (desde fila 2 que es el header)
    ws.auto_filter.ref = f"A2:M{len(datos) + 2}"

    # Guardar
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Crear backup del archivo existente antes de sobrescribir
    backup_dir = OUTPUT_DIR / "backups"
    if OUTPUT_FILE.exists():
        try:
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = backup_dir / f"control_ejecucion_{timestamp}.xlsx"

            # Copiar archivo actual a backups
            import shutil
            shutil.copy2(str(OUTPUT_FILE), str(backup_path))
            print(f"  Backup creado: backups/{backup_path.name}")
        except Exception as e:
            print(f"  Aviso: No se pudo crear backup: {e}")

    # Guardar archivo principal
    try:
        wb.save(str(OUTPUT_FILE))
        print(f"  Archivo actualizado: {OUTPUT_FILE.name}")
    except PermissionError:
        print(f"  Error: No se pudo guardar {OUTPUT_FILE.name} (archivo abierto)")
        print("  Cierre el archivo Excel e intente nuevamente.")

    return str(OUTPUT_FILE)


def generar_resumen(datos):
    """Genera un resumen de los datos procesados"""
    print("\n" + "="*60)
    print("RESUMEN DEL ARCHIVO DE CONTROL")
    print("="*60)

    total_hus = len(datos)

    # Por estado de certificacion
    estados = {}
    for d in datos:
        estado = d['Estado Certificacion']
        estados[estado] = estados.get(estado, 0) + 1

    print(f"\nTotal HUs: {total_hus}")
    print("\nPor Estado de Certificacion:")
    for estado, count in estados.items():
        pct = (count / total_hus) * 100 if total_hus > 0 else 0
        print(f"  {estado}: {count} ({pct:.1f}%)")

    # Por categoria
    categorias = {}
    for d in datos:
        cat = d['Categoria'] if d['Categoria'] else "(Sin categoria)"
        categorias[cat] = categorias.get(cat, 0) + 1

    print("\nPor Categoria:")
    for cat, count in categorias.items():
        print(f"  {cat}: {count}")

    # Por tester
    testers = {}
    for d in datos:
        tester = d['Tester Asignado'] if d['Tester Asignado'] else "(Sin asignar)"
        testers[tester] = testers.get(tester, 0) + 1

    print("\nPor Tester:")
    for tester, count in sorted(testers.items(), key=lambda x: x[1], reverse=True):
        print(f"  {tester}: {count}")

    # Totales de TCs
    total_tcs = sum(d['Total Test Cases'] for d in datos)
    total_completos = sum(d['Test Cases Completos'] for d in datos)
    avance_global = (total_completos / total_tcs * 100) if total_tcs > 0 else 0

    print(f"\nTest Cases:")
    print(f"  Total: {total_tcs}")
    print(f"  Completos: {total_completos}")
    print(f"  Pendientes: {total_tcs - total_completos}")
    print(f"  Avance Global: {avance_global:.1f}%")


def main():
    """Funcion principal"""
    print("="*60)
    print("GENERADOR DE ARCHIVO DE CONTROL DE EJECUCION")
    print("="*60)
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Verificar que el archivo no este abierto antes de proceder
    if not verificar_archivo_cerrado():
        return  # Usuario cancelo

    # Cargar datos manuales existentes (Estado Subida, Observaciones)
    datos_manuales = cargar_datos_manuales_existentes()

    # Cargar CSV
    df = cargar_csv()
    if df is None:
        return

    # Construir mapa HU -> TCs
    mapa = construir_mapa_hu_tcs(df)

    # Generar datos de control (conservando datos manuales)
    datos = generar_datos_control(mapa, datos_manuales)

    # Crear archivo Excel
    print("\nGenerando archivo Excel...")
    output_path = crear_archivo_excel(datos)

    # Mostrar resumen
    generar_resumen(datos)

    print("\n" + "="*60)
    print(f"Archivo generado: {output_path}")
    print(f"Archivo principal: {OUTPUT_FILE}")
    print("="*60)


if __name__ == "__main__":
    main()
